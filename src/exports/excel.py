"""Excel export functionality.

Generates Ford-themed Excel reports for:
- Standings by age group/division
- Recommendations with explanations
- SoS analysis
"""

from __future__ import annotations

import io
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

from src.grading.metrics import calculate_team_metrics
from src.grading.strength_of_schedule import calculate_sos
from src.models.game_result import TeamSeason
from src.models.recommendation import Recommendation, RecommendationType

# Ford theme colors
FORD_BLUE = "00095B"
ACCENT_BLUE = "1C69D4"
LIGHT_BLUE = "E8F1FC"
SUCCESS_GREEN = "198754"
WARNING_YELLOW = "FFC107"
ERROR_RED = "DC3545"
WHITE = "FFFFFF"


def export_standings_excel(
    teams: list[TeamSeason],
    recommendations: dict[str, Recommendation] | None = None,
    output_path: Path | None = None,
) -> bytes:
    """Export standings to Ford-themed Excel file.

    Args:
        teams: Teams to include in standings.
        recommendations: Optional recommendations dict.
        output_path: Optional path to save file (otherwise returns bytes).

    Returns:
        Excel file as bytes.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Standings"

    # Build data
    rows = []
    for team in teams:
        metrics = calculate_team_metrics(team)
        sos = calculate_sos(team)

        rec = recommendations.get(team.team_name) if recommendations else None

        row = {
            "Team": team.team_name,
            "Club": team.club_name or "",
            "Grade": team.assigned_grade.value
            if hasattr(team.assigned_grade, "value")
            else str(team.assigned_grade)
            if team.assigned_grade
            else "",
            "Division": team.division,
            "W": metrics.wins,
            "L": metrics.losses,
            "Win%": metrics.win_rate,
            "PF": metrics.points_for,
            "PA": metrics.points_against,
            "+/-": metrics.total_margin,
            "Blowout W": metrics.blowout_wins,
            "Blowout L": metrics.blowout_losses,
            "SoS Score": sos.sos_score,
            "Schedule": sos.schedule_difficulty,
            "Status": (
                rec.recommendation_type.value
                if hasattr(rec.recommendation_type, "value")
                else str(rec.recommendation_type)
            )
            if rec
            else "N/A",
        }
        rows.append(row)

    df = pd.DataFrame(rows)
    df = df.sort_values("+/-", ascending=False)

    # Write headers with Ford Blue styling
    header_fill = PatternFill(start_color=FORD_BLUE, end_color=FORD_BLUE, fill_type="solid")
    header_font = Font(bold=True, color=WHITE)
    header_alignment = Alignment(horizontal="center", vertical="center")

    for col_idx, header in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Write data with alternating row colors
    light_fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")

    for row_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if row_idx % 2 == 0:
                cell.fill = light_fill

            # Color code status column
            if col_idx == len(df.columns):  # Status column
                if value == "promote":
                    cell.fill = PatternFill(
                        start_color=SUCCESS_GREEN, end_color=SUCCESS_GREEN, fill_type="solid"
                    )
                    cell.font = Font(color=WHITE)
                elif value == "demote":
                    cell.fill = PatternFill(
                        start_color=ERROR_RED, end_color=ERROR_RED, fill_type="solid"
                    )
                    cell.font = Font(color=WHITE)
                elif value == "review_needed":
                    cell.fill = PatternFill(
                        start_color=WARNING_YELLOW, end_color=WARNING_YELLOW, fill_type="solid"
                    )

    # Auto-fit columns
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except TypeError:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save to bytes
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    if output_path:
        with open(output_path, "wb") as f:
            f.write(buffer.getvalue())

    return buffer.getvalue()


def export_recommendations_excel(
    recommendations: list[Recommendation],
    output_path: Path | None = None,
) -> bytes:
    """Export recommendations to Ford-themed Excel file.

    Args:
        recommendations: Recommendations to export.
        output_path: Optional path to save file.

    Returns:
        Excel file as bytes.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Recommendations"

    # Headers
    headers = [
        "Team",
        "Current Grade",
        "Recommended Grade",
        "Recommendation",
        "Confidence",
        "Explanation",
        "Evidence",
        "Concerns",
        "SoS Note",
    ]

    header_fill = PatternFill(start_color=FORD_BLUE, end_color=FORD_BLUE, fill_type="solid")
    header_font = Font(bold=True, color=WHITE)

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font

    # Data rows
    for row_idx, rec in enumerate(recommendations, 2):
        ws.cell(row=row_idx, column=1, value=rec.team_name)
        ws.cell(row=row_idx, column=2, value=rec.current_grade or "N/A")
        ws.cell(row=row_idx, column=3, value=rec.recommended_grade or "No change")
        rec_type = (
            rec.recommendation_type.value
            if hasattr(rec.recommendation_type, "value")
            else str(rec.recommendation_type)
        )
        conf_val = rec.confidence.value if hasattr(rec.confidence, "value") else str(rec.confidence)
        ws.cell(row=row_idx, column=4, value=rec_type)
        ws.cell(row=row_idx, column=5, value=conf_val)
        ws.cell(row=row_idx, column=6, value=rec.explanation)
        ws.cell(row=row_idx, column=7, value="; ".join(rec.evidence))
        ws.cell(row=row_idx, column=8, value="; ".join(rec.concerns))
        ws.cell(row=row_idx, column=9, value=rec.strength_of_schedule_note or "")

        # Color code recommendation column
        rec_cell = ws.cell(row=row_idx, column=4)
        if rec.recommendation_type == RecommendationType.PROMOTE:
            rec_cell.fill = PatternFill(
                start_color=SUCCESS_GREEN, end_color=SUCCESS_GREEN, fill_type="solid"
            )
            rec_cell.font = Font(color=WHITE)
        elif rec.recommendation_type == RecommendationType.DEMOTE:
            rec_cell.fill = PatternFill(
                start_color=ERROR_RED, end_color=ERROR_RED, fill_type="solid"
            )
            rec_cell.font = Font(color=WHITE)
        elif rec.recommendation_type == RecommendationType.REVIEW_NEEDED:
            rec_cell.fill = PatternFill(
                start_color=WARNING_YELLOW, end_color=WARNING_YELLOW, fill_type="solid"
            )

    # Auto-fit columns
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except TypeError:
                pass
        adjusted_width = min(max_length + 2, 60)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save to bytes
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    if output_path:
        with open(output_path, "wb") as f:
            f.write(buffer.getvalue())

    return buffer.getvalue()


def export_audit_log_excel(
    overrides: list,
    output_path: Path | None = None,
) -> bytes:
    """Export audit log to Excel.

    Args:
        overrides: List of Override objects.
        output_path: Optional path to save file.

    Returns:
        Excel file as bytes.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Audit Log"

    headers = [
        "Date",
        "Team",
        "Original Recommendation",
        "Admin Decision",
        "Final Grade",
        "Reason",
        "Admin ID",
        "Season",
        "Round",
    ]

    header_fill = PatternFill(start_color=FORD_BLUE, end_color=FORD_BLUE, fill_type="solid")
    header_font = Font(bold=True, color=WHITE)

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font

    for row_idx, override in enumerate(overrides, 2):
        ws.cell(row=row_idx, column=1, value=override.timestamp.strftime("%Y-%m-%d %H:%M"))
        ws.cell(row=row_idx, column=2, value=override.team_name)
        ws.cell(row=row_idx, column=3, value=override.original_recommendation)
        ws.cell(row=row_idx, column=4, value=override.admin_decision)
        ws.cell(row=row_idx, column=5, value=override.final_grade or "N/A")
        ws.cell(row=row_idx, column=6, value=override.reason)
        ws.cell(row=row_idx, column=7, value=override.admin_id or "N/A")
        ws.cell(row=row_idx, column=8, value=override.season or "N/A")
        ws.cell(row=row_idx, column=9, value=override.round_num or "N/A")

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    if output_path:
        with open(output_path, "wb") as f:
            f.write(buffer.getvalue())

    return buffer.getvalue()


def export_matchups_excel(
    matchups: list,
    round_num: int,
    output_path: Path | None = None,
) -> bytes:
    """Export matchup recommendations to Ford-themed Excel file.

    Creates one tab per grade/age combination (e.g., "U12 Boys A").

    Args:
        matchups: List of MatchupRecommendation objects.
        round_num: Round number for the matchups.
        output_path: Optional path to save file.

    Returns:
        Excel file as bytes.
    """
    import re
    from collections import Counter, defaultdict

    from src.models.matchup import MatchupType

    wb = Workbook()

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # Group matchups by age group and grade
    grouped: dict[str, list] = defaultdict(list)
    for matchup in matchups:
        # Determine the primary grade for this matchup
        grade_a = (
            matchup.team_a_grade.value
            if hasattr(matchup.team_a_grade, "value")
            else str(matchup.team_a_grade)
            if matchup.team_a_grade
            else "Unknown"
        )
        grade_b = (
            matchup.team_b_grade.value
            if hasattr(matchup.team_b_grade, "value")
            else str(matchup.team_b_grade)
            if matchup.team_b_grade
            else "Unknown"
        )

        # Use the higher grade as the primary for cross-grade matchups
        primary_grade = grade_a if grade_a <= grade_b else grade_b

        # Extract age group from team name (assumes format like "Jets U12 Boys 1")
        age_match = re.search(r"U(\d+)", matchup.team_a_name)
        age_group = f"U{age_match.group(1)}" if age_match else "Unknown"

        gender_match = re.search(r"(Boys|Girls)", matchup.team_a_name, re.IGNORECASE)
        gender = gender_match.group(1).title() if gender_match else "Unknown"

        tab_name = f"{age_group} {gender} {primary_grade}"
        grouped[tab_name].append(matchup)

    # Create tabs for each group
    for tab_name in sorted(grouped.keys()):
        group_matchups = grouped[tab_name]

        # Truncate tab name if too long (Excel limit is 31 chars)
        safe_tab_name = tab_name[:31]
        ws = wb.create_sheet(title=safe_tab_name)

        # Write title
        ws.cell(row=1, column=1, value=f"Round {round_num} Matchups - {tab_name}")
        ws.cell(row=1, column=1).font = Font(bold=True, size=14, color=FORD_BLUE)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)

        # Headers
        headers = ["Team A", "Team A Grade", "Team B", "Team B Grade", "Type", "Notes"]

        header_fill = PatternFill(start_color=FORD_BLUE, end_color=FORD_BLUE, fill_type="solid")
        header_font = Font(bold=True, color=WHITE)

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        light_fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
        cross_grade_fill = PatternFill(
            start_color=WARNING_YELLOW, end_color=WARNING_YELLOW, fill_type="solid"
        )

        for row_idx, matchup in enumerate(group_matchups, 4):
            grade_a = (
                matchup.team_a_grade.value
                if hasattr(matchup.team_a_grade, "value")
                else str(matchup.team_a_grade)
                if matchup.team_a_grade
                else "?"
            )
            grade_b = (
                matchup.team_b_grade.value
                if hasattr(matchup.team_b_grade, "value")
                else str(matchup.team_b_grade)
                if matchup.team_b_grade
                else "?"
            )
            type_str = (
                matchup.matchup_type.value
                if isinstance(matchup.matchup_type, MatchupType)
                else matchup.matchup_type
            )
            type_display = type_str.replace("_", " ").title()

            notes = ""
            if matchup.justification:
                notes = (
                    matchup.justification[:100] + "..."
                    if len(matchup.justification) > 100
                    else matchup.justification
                )

            ws.cell(row=row_idx, column=1, value=matchup.team_a_name)
            ws.cell(row=row_idx, column=2, value=grade_a)
            ws.cell(row=row_idx, column=3, value=matchup.team_b_name)
            ws.cell(row=row_idx, column=4, value=grade_b)
            ws.cell(row=row_idx, column=5, value=type_display)
            ws.cell(row=row_idx, column=6, value=notes)

            # Highlight cross-grade matchups
            is_cross_grade = type_str != "same_grade"
            for col_idx in range(1, 7):
                cell = ws.cell(row=row_idx, column=col_idx)
                if is_cross_grade:
                    cell.fill = cross_grade_fill
                elif row_idx % 2 == 0:
                    cell.fill = light_fill

        # Auto-fit columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except TypeError:
                    pass
            adjusted_width = min(max_length + 2, 40)
            ws.column_dimensions[column_letter].width = adjusted_width

    # Create summary sheet
    ws_summary = wb.create_sheet(title="Summary", index=0)
    ws_summary.cell(row=1, column=1, value=f"Round {round_num} Matchups Summary")
    ws_summary.cell(row=1, column=1).font = Font(bold=True, size=14, color=FORD_BLUE)

    ws_summary.cell(row=3, column=1, value="Total Matchups:")
    ws_summary.cell(row=3, column=2, value=len(matchups))
    ws_summary.cell(row=3, column=1).font = Font(bold=True)

    # Count by type
    type_counts = Counter(
        m.matchup_type.value if isinstance(m.matchup_type, MatchupType) else m.matchup_type
        for m in matchups
    )

    ws_summary.cell(row=5, column=1, value="Matchup Types:")
    ws_summary.cell(row=5, column=1).font = Font(bold=True)
    summary_row = 6
    for mtype, count in sorted(type_counts.items()):
        ws_summary.cell(row=summary_row, column=1, value=mtype.replace("_", " ").title())
        ws_summary.cell(row=summary_row, column=2, value=count)
        summary_row += 1

    # Tabs info
    ws_summary.cell(row=summary_row + 1, column=1, value="Tabs:")
    ws_summary.cell(row=summary_row + 1, column=1).font = Font(bold=True)
    summary_row += 2
    for tab_name, group_matchups in sorted(grouped.items()):
        ws_summary.cell(row=summary_row, column=1, value=tab_name)
        ws_summary.cell(row=summary_row, column=2, value=len(group_matchups))
        summary_row += 1

    # Save to bytes
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    if output_path:
        with open(output_path, "wb") as f:
            f.write(buffer.getvalue())

    return buffer.getvalue()
