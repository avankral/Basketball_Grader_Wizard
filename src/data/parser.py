"""Excel parser for DVBA Club Grading Book format.

This module parses the wide-format Excel files used by DVBA for basketball
grading, transforming them into structured data for analysis.

The parser handles:
- Multi-sheet workbooks (one sheet per gender/age/division)
- Wide format with repeating round columns
- Score string parsing ("31 - 56" → (31, 56))
- Cell color detection for variance classification
- DNP (Did Not Play) handling
"""

from __future__ import annotations

import contextlib
import re
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.cell import Cell
from openpyxl.styles import PatternFill

from src.models.game_result import (
    GameResult,
    Grade,
    ResultType,
    SheetInfo,
    TeamSeason,
    VarianceClass,
)

# Color mapping for variance classification (approximate RGB values)
# These may need adjustment based on actual Excel colors used
VARIANCE_COLOR_MAP = {
    # Blue shades (blowout win)
    "0000FF": VarianceClass.BLOWOUT_WIN,
    "0066CC": VarianceClass.BLOWOUT_WIN,
    "4472C4": VarianceClass.BLOWOUT_WIN,
    "5B9BD5": VarianceClass.BLOWOUT_WIN,
    # Green shades (dominant win)
    "00FF00": VarianceClass.DOMINANT_WIN,
    "00B050": VarianceClass.DOMINANT_WIN,
    "92D050": VarianceClass.DOMINANT_WIN,
    "70AD47": VarianceClass.DOMINANT_WIN,
    "A9D08E": VarianceClass.DOMINANT_WIN,
    # Red shades (bad loss - dark red)
    "FF0000": VarianceClass.BAD_LOSS,
    "C00000": VarianceClass.BAD_LOSS,
    # Light red / pink (concerning loss)
    "FF9999": VarianceClass.CONCERNING_LOSS,
    "FFCCCC": VarianceClass.CONCERNING_LOSS,
    "F4B084": VarianceClass.CONCERNING_LOSS,
    "FCE4D6": VarianceClass.CONCERNING_LOSS,
}


def get_cell_color(cell: Cell) -> str | None:
    """Extract fill color from an Excel cell.

    Args:
        cell: openpyxl Cell object.

    Returns:
        Hex color string (e.g., "FF0000") or None if no fill.
    """
    if cell.fill is None:
        return None

    fill: PatternFill = cell.fill
    if fill.patternType is None or fill.patternType == "none":
        return None

    # Get foreground color (the fill color)
    fg_color = fill.fgColor
    if fg_color is None:
        return None

    # Handle different color types
    if fg_color.type == "rgb" and fg_color.rgb:
        # Remove alpha channel if present (ARGB -> RGB)
        rgb = fg_color.rgb
        if len(rgb) == 8:
            rgb = rgb[2:]  # Skip first 2 chars (alpha)
        return rgb.upper()

    if fg_color.type == "indexed" and fg_color.indexed is not None:
        # Indexed colors - map common ones
        # This is a simplified mapping; may need expansion
        indexed_map = {
            0: None,  # No fill
            1: None,
            2: "FF0000",  # Red
            3: "00FF00",  # Green
            4: "0000FF",  # Blue
            5: "FFFF00",  # Yellow
        }
        return indexed_map.get(fg_color.indexed)

    return None


def classify_variance(cell: Cell) -> VarianceClass:
    """Classify game variance based on cell fill color.

    Args:
        cell: openpyxl Cell object.

    Returns:
        VarianceClass indicating the type of result.
    """
    color = get_cell_color(cell)
    if color is None:
        return VarianceClass.COMPETITIVE

    # Exact match
    if color in VARIANCE_COLOR_MAP:
        return VARIANCE_COLOR_MAP[color]

    # Fuzzy match by color family
    # This handles slight variations in color codes
    r = int(color[0:2], 16) if len(color) >= 2 else 0
    g = int(color[2:4], 16) if len(color) >= 4 else 0
    b = int(color[4:6], 16) if len(color) >= 6 else 0

    # Blue dominant = blowout win
    if b > 150 and b > r and b > g:
        return VarianceClass.BLOWOUT_WIN

    # Green dominant = dominant win
    if g > 150 and g > r and g > b:
        return VarianceClass.DOMINANT_WIN

    # Red dominant = loss
    if r > 150 and r > g:
        if g < 100:
            return VarianceClass.BAD_LOSS
        else:
            return VarianceClass.CONCERNING_LOSS

    return VarianceClass.COMPETITIVE


def parse_score_string(score_str: str) -> tuple[int, int]:
    """Parse a score string into (score_for, score_against).

    Args:
        score_str: Score string like "31 - 56" or "31-56"

    Returns:
        Tuple of (score_for, score_against).

    Raises:
        ValueError: If score string cannot be parsed.
    """
    if not score_str or not isinstance(score_str, str):
        raise ValueError(f"Invalid score string: {score_str}")

    # Remove extra whitespace and normalize separators
    normalized = score_str.strip().replace("–", "-").replace("—", "-")

    # Pattern: number - number
    pattern = r"(\d+)\s*-\s*(\d+)"
    match = re.search(pattern, normalized)

    if not match:
        raise ValueError(f"Could not parse score: {score_str}")

    return int(match.group(1)), int(match.group(2))


class GradingBookParser:
    """Parser for DVBA Club Grading Book Excel files.

    Transforms wide-format Excel data into structured TeamSeason objects
    with embedded GameResult records.
    """

    def __init__(self, file_path: Path | str):
        """Initialize parser with path to Excel file.

        Args:
            file_path: Path to the .xlsx file.
        """
        self.file_path = Path(file_path)
        if not self.file_path.exists():
            raise FileNotFoundError(f"Excel file not found: {self.file_path}")

        self._workbook = None
        self._team_grades: dict[str, Grade] = {}  # Cache team -> grade mapping

    def _load_workbook(self) -> None:
        """Load the Excel workbook if not already loaded."""
        if self._workbook is None:
            self._workbook = load_workbook(self.file_path, data_only=False)

    def get_sheet_names(self) -> list[str]:
        """Get all sheet names from the workbook.

        Returns:
            List of sheet names.
        """
        self._load_workbook()
        return self._workbook.sheetnames

    def get_valid_sheets(self) -> list[SheetInfo]:
        """Get sheet info for all valid grading sheets.

        Returns:
            List of SheetInfo for sheets matching the pattern {Gender}{Age}{Div}.
        """
        valid_sheets = []
        for name in self.get_sheet_names():
            info = SheetInfo.from_sheet_name(name)
            if info is not None:
                valid_sheets.append(info)
        return valid_sheets

    def _detect_round_columns(self, df: pd.DataFrame) -> list[int]:
        """Detect which columns contain round data.

        Looks for column headers starting with "Rnd" or round number patterns.

        Args:
            df: DataFrame with headers as first row.

        Returns:
            List of column indices where round blocks start.
        """
        round_cols = []
        for idx, col in enumerate(df.columns):
            col_str = str(col).strip()
            # Match "Rnd 1", "Rnd 2", or similar patterns
            if re.match(r"Rnd\s*\d+", col_str, re.IGNORECASE):
                round_cols.append(idx)
        return round_cols

    def _build_team_grade_cache(self) -> None:
        """Build a cache mapping team names to their assigned grades.

        This enables looking up opponent grades when parsing game results.
        """
        self._load_workbook()
        self._team_grades.clear()

        for sheet_name in self.get_sheet_names():
            info = SheetInfo.from_sheet_name(sheet_name)
            if info is None:
                continue

            ws = self._workbook[sheet_name]

            # Find the team and grade columns
            # Typically: Column B = Team, Column C = Grade
            for row in ws.iter_rows(min_row=2, max_col=10):
                team_cell = row[1] if len(row) > 1 else None  # Column B (0-indexed)
                grade_cell = row[2] if len(row) > 2 else None  # Column C

                if team_cell and team_cell.value:
                    team_name = str(team_cell.value).strip()
                    if team_name and grade_cell and grade_cell.value:
                        grade = Grade.from_string(str(grade_cell.value).strip())
                        if grade:
                            self._team_grades[team_name] = grade

    def _get_opponent_grade(self, opponent_name: str) -> Grade | None:
        """Look up the assigned grade for an opponent.

        Args:
            opponent_name: Full opponent team name.

        Returns:
            Grade if found, None otherwise.
        """
        if not self._team_grades:
            self._build_team_grade_cache()

        # Try exact match first
        if opponent_name in self._team_grades:
            return self._team_grades[opponent_name]

        # Try normalized match (strip extra whitespace)
        normalized = " ".join(opponent_name.split())
        if normalized in self._team_grades:
            return self._team_grades[normalized]

        # Try partial match for slight variations
        for cached_name, grade in self._team_grades.items():
            if cached_name.lower() == opponent_name.lower():
                return grade

        return None

    def parse_sheet(self, sheet_name: str) -> list[TeamSeason]:
        """Parse a single sheet into TeamSeason objects.

        Args:
            sheet_name: Name of the sheet to parse.

        Returns:
            List of TeamSeason objects, one per team in the sheet.
        """
        self._load_workbook()

        info = SheetInfo.from_sheet_name(sheet_name)
        if info is None:
            return []

        ws = self._workbook[sheet_name]

        # Read sheet into pandas for easier manipulation
        data = list(ws.values)
        if len(data) < 2:
            return []

        # First row is headers
        headers = [str(h) if h else f"col_{i}" for i, h in enumerate(data[0])]
        df = pd.DataFrame(data[1:], columns=headers)

        # Build team grade cache if not done
        if not self._team_grades:
            self._build_team_grade_cache()

        teams = []

        # Process each team row
        for row_idx, row in df.iterrows():
            # Extract team info (first few columns)
            team_name = str(row.iloc[1]).strip() if pd.notna(row.iloc[1]) else None
            if not team_name:
                continue

            grade_str = str(row.iloc[2]).strip() if pd.notna(row.iloc[2]) else None
            assigned_grade = Grade.from_string(grade_str) if grade_str else None

            rank = None
            if pd.notna(row.iloc[3]):
                with contextlib.suppress(ValueError, TypeError):
                    rank = int(float(row.iloc[3]))

            club_name = TeamSeason.extract_club_name(team_name)

            team = TeamSeason(
                team_name=team_name,
                club_name=club_name,
                gender=info.gender,
                age_group=info.age_group,
                division=info.division,
                assigned_grade=assigned_grade,
                current_rank=rank,
                games=[],
                sheet_name=sheet_name,
            )

            # Parse round data
            # Round blocks typically start at column D (index 3) and repeat
            # Each round has ~6-7 columns: Result, Score, Margin, Opponent, Sheet, Description
            round_num = 1
            col_idx = 4  # Start after Rank column

            while col_idx < len(row) - 5:  # Need at least 6 columns for a round
                # Get cell values
                result_val = row.iloc[col_idx] if pd.notna(row.iloc[col_idx]) else None
                score_val = (
                    row.iloc[col_idx + 1]
                    if col_idx + 1 < len(row) and pd.notna(row.iloc[col_idx + 1])
                    else None
                )
                margin_val = (
                    row.iloc[col_idx + 2]
                    if col_idx + 2 < len(row) and pd.notna(row.iloc[col_idx + 2])
                    else None
                )
                opponent_val = (
                    row.iloc[col_idx + 3]
                    if col_idx + 3 < len(row) and pd.notna(row.iloc[col_idx + 3])
                    else None
                )
                sheet_val = (
                    row.iloc[col_idx + 4]
                    if col_idx + 4 < len(row) and pd.notna(row.iloc[col_idx + 4])
                    else None
                )

                # Check if this looks like a round block
                if not result_val:
                    col_idx += 1
                    continue

                result_str = str(result_val).strip()
                result_type = ResultType.from_string(result_str)

                if result_type is None:
                    # Check if this is a header row for round
                    if result_str.lower().startswith("rnd"):
                        # Skip this and try again
                        col_idx += 1
                        continue
                    col_idx += 1
                    continue

                # Skip DNP entries
                if result_type == ResultType.DNP:
                    col_idx += 6  # Move to next round block
                    round_num += 1
                    continue

                # Parse score
                try:
                    if score_val:
                        score_for, score_against = parse_score_string(str(score_val))
                    else:
                        col_idx += 6
                        round_num += 1
                        continue
                except ValueError:
                    col_idx += 6
                    round_num += 1
                    continue

                # Get margin (use calculated if not provided)
                if margin_val:
                    try:
                        margin = int(float(margin_val))
                    except (ValueError, TypeError):
                        margin = score_for - score_against
                else:
                    margin = score_for - score_against

                # Get opponent info
                opponent_name = str(opponent_val).strip() if opponent_val else "Unknown"
                opponent_sheet = str(sheet_val).strip() if sheet_val else None
                opponent_grade = self._get_opponent_grade(opponent_name)

                # Get variance class from cell color
                # Note: We need to access the original cell for color info
                ws_row = row_idx + 2  # +1 for header, +1 for 1-indexed
                result_cell = ws.cell(row=ws_row, column=col_idx + 1)  # +1 for 1-indexed
                variance = classify_variance(result_cell)

                game = GameResult(
                    team_name=team_name,
                    opponent_name=opponent_name,
                    opponent_grade=opponent_grade,
                    score_for=score_for,
                    score_against=score_against,
                    margin=margin,
                    result=result_type,
                    round_num=round_num,
                    variance_class=variance,
                    opponent_sheet=opponent_sheet,
                )
                team.games.append(game)

                col_idx += 6  # Move to next round block
                round_num += 1

            if team.games:  # Only add teams with actual games
                teams.append(team)

        return teams

    def parse_all(self) -> list[TeamSeason]:
        """Parse all valid sheets in the workbook.

        Returns:
            List of all TeamSeason objects from all sheets.
        """
        self._load_workbook()
        self._build_team_grade_cache()

        all_teams = []
        for info in self.get_valid_sheets():
            teams = self.parse_sheet(info.sheet_name)
            all_teams.extend(teams)

        return all_teams

    def to_dataframe(self) -> pd.DataFrame:
        """Parse all sheets and return as a single long-format DataFrame.

        Returns:
            DataFrame with one row per game result.
        """
        teams = self.parse_all()

        records = []
        for team in teams:
            for game in team.games:
                # Handle both enum and string values (Pydantic use_enum_values)
                def get_value(val):
                    if val is None:
                        return None
                    return val.value if hasattr(val, "value") else val

                records.append(
                    {
                        "team_name": team.team_name,
                        "club_name": team.club_name,
                        "gender": get_value(team.gender),
                        "age_group": team.age_group,
                        "division": team.division,
                        "assigned_grade": get_value(team.assigned_grade),
                        "sheet_name": team.sheet_name,
                        "round_num": game.round_num,
                        "opponent_name": game.opponent_name,
                        "opponent_grade": get_value(game.opponent_grade),
                        "score_for": game.score_for,
                        "score_against": game.score_against,
                        "margin": game.margin,
                        "result": get_value(game.result),
                        "variance_class": get_value(game.variance_class),
                        "opponent_sheet": game.opponent_sheet,
                        "is_blowout": game.is_blowout,
                        "is_close_game": game.is_close_game,
                    }
                )

        return pd.DataFrame(records)


def parse_grading_book(file_path: Path | str) -> pd.DataFrame:
    """Convenience function to parse a grading book to DataFrame.

    Args:
        file_path: Path to the .xlsx file.

    Returns:
        Long-format DataFrame with one row per game.
    """
    parser = GradingBookParser(file_path)
    return parser.to_dataframe()
